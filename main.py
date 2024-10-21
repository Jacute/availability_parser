from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait

from bs4 import BeautifulSoup

from webdriver_manager.chrome import ChromeDriverManager

from datetime import datetime

from config.config import *

import urllib3
from urllib3.exceptions import InsecureRequestWarning
from requests.exceptions import RequestException
from time import sleep
import requests
import json
import re
import os
import sys
import traceback
import argparse


USER_AGENT = 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:129.0) Gecko/20100101 Firefox/129.0'


class Parser:
    def __init__(self):
        self.result = {}
        parser = argparse.ArgumentParser(description='Process some integers.')
        parser.add_argument('--headless', action='store_true', help='headless')
        self.args = parser.parse_args()
        if self.args.headless:
            self.driver = self.get_driver(True)
        else:
            self.driver = self.get_driver(False)

    def get_driver(self, headless):
        try:
            options = webdriver.ChromeOptions()
            if headless:
                options.add_argument('--headless')
                options.add_argument('--disable-gpu')

            options.add_argument('--log-level=3')
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)

            options.add_argument(
                "user-agent=Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:84.0) Gecko/20100101 Firefox/84.0")

            # options.add_argument('--disable-dev-shm-usage')
            # options.add_argument('--no-sandbox')
            service = Service(WEBDRIVER_PATH)
            driver = webdriver.Chrome(
                # service=Service(ChromeDriverManager().install()),
                service=service,
                options=options
            )
            driver.set_window_size(1920, 1080)
            driver.implicitly_wait(30)

            self.wait = WebDriverWait(driver, 30)

            return driver
        except Exception as e:
            print('Неудачная настройка браузера!')
            print(traceback.format_exc())
            print(input('Нажмите ENTER, чтобы закрыть эту программу'))
            sys.exit()

    def parseOne(self, prefix, article, size, length):
        if prefix == 'H&M':
            url = f'https://www2.hm.com/pl_pl/productpage.{article}.html'
            self.driver.get(url)

            pricesText = self.driver.find_element(By.ID, 'product-price').text.replace(',', '.').replace('PLN',
                                                                                                        '').strip()
            prices = re.findall(r'[0-9 ]+\.\d{2}', pricesText)
            if len(prices) == 0:
                prices = re.findall(r'\d+', prices)
            if len(prices) == 1:
                price = self.get_hm_price(prices[0].replace(' ', ''))
                sale_price = ''
            else:
                if 'Cena dla Klubowiczów' in pricesText:
                    price = self.get_hm_price(prices[0].replace(' ', ''))
                    sale_price = self.get_hm_price(prices[1].replace(' ', ''))
                else:
                    price = self.get_hm_price(prices[1].replace(' ', ''))
                    sale_price = self.get_hm_price(prices[0].replace(' ', ''))
            if size:
                sizes = self.driver.find_elements(By.XPATH, '//hm-size-selector/ul/li/div/label')
                for elem in sizes:
                    new_article = prefix + '_' + article + '_' + elem.text.split('\n')[0].strip()
                    if 'Zostało tylko kilka sztuk!' in elem.text:
                        self.result[new_article] = [self.AVIABLE_HM["few_items"], price, sale_price]
                    elif elem.get_attribute('aria-disabled') == 'true':
                        self.result[new_article] = [self.AVIABLE_HM["no_aviable"], price, sale_price]
                    else:
                        self.result[new_article] = [self.AVIABLE_HM["aviable"], price, sale_price]
            else:  # Для сумок
                new_article = prefix + '_' + article
                btn = self.driver.find_element(By.CLASS_NAME, 'item.button.fluid')
                if 'Dodaj' not in btn.text:
                    self.result[new_article] = [self.AVIABLE_HM["no_aviable"], price, sale_price]
                else:
                    self.result[new_article] = [self.AVIABLE_HM["aviable"], price, sale_price]
        elif prefix == 'COS':
            url = f'https://www.cos.com/en_eur/women/womenswear/t-shirts/product.the-full-volume-t-shirt-green.{article}.html'
            self.driver.get(url)

            eur_prices = self.driver.find_element(By.XPATH, '//div[@class="price parbase"]').text.replace('€',
                                                                                                            '').replace(
                ',', '.').strip()
            eur_prices = re.findall(r'\d+\.\d{2}', eur_prices)
            if len(eur_prices) == 1:
                price = self.get_cos_price(eur_prices[0])
                sale_price = ''
            else:
                price = self.get_cos_price(eur_prices[0])
                sale_price = self.get_cos_price(eur_prices[1])
            if size:
                sizes = self.driver.find_elements(By.XPATH, '//div[@class="size-container"]/button')
                for elem in sizes:
                    new_article = prefix + '_' + article + '_' + elem.text.split('\n')[0]
                    if 'out-of-stock' in elem.get_attribute('class'):
                        self.result[new_article] = [self.AVIABLE_COS['no_aviable'], price, sale_price]
                    elif 'low-in-stock-size' in elem.get_attribute('class'):
                        self.result[new_article] = [self.AVIABLE_COS["few_items"], price, sale_price]
                    else:
                        self.result[new_article] = [self.AVIABLE_COS["aviable"], price, sale_price]
            else:
                pass
        elif prefix == 'ASOS':
            r = self.make_request(f'https://www.asos.com/search/?q={article}', headers={'User-Agent': USER_AGENT}, cookies={'browseCountry': 'EE', 'browseCurrency': 'EUR'})
            
            productData = re.search(r'window.asos.pdp.config.product = {(.+)};', r.text)
            productData = json.loads('{' + productData.group(1) + '}')
            id = productData["id"]
            idSize = {}
            for variant in productData["variants"]:
                idSize[variant["variantId"]] = variant["size"]
            price_response = self.make_request(f'https://www.asos.com/api/product/catalogue/v4/stockprice?productIds={id}&store=PL&currency=EUR&keyStoreDataversion=11a1qu9-40&country=PL')
            price_info = price_response.json()
            for variant in price_info[0]["variants"]:
                id = variant["id"]
                is_in_stock = variant["isInStock"]
                price = self.get_cos_price(variant["price"]["previous"]["value"])
                sale_price = self.get_cos_price(variant["price"]["current"]["value"])
                if price == sale_price:
                    sale_price = None
                size = idSize[id]
                if size == 'No Size':
                    size = ''
                
                new_article = prefix + '_' + article + '_' + size
                print(new_article)
                if is_in_stock:
                    self.result[new_article] = [self.AVIABLE_ASOS["aviable"], price, sale_price]
                else:
                    self.result[new_article] = [self.AVIABLE_ASOS["no_aviable"], price, sale_price]
        elif prefix == 'UNIQLO':
            host = 'https://www.uniqlo.com'
            
            res = self.make_request(f'{host}/eu/en_AT/search?q={article}', headers={'User-Agent': USER_AGENT})
            soup = BeautifulSoup(res.text, 'lxml')

            colors = [j.get('data-replaceurl') for j in
                        soup.find_all('button', {'data-attr-name': "color"})]
            for j in colors:
                res = self.make_request(j, headers={'User-Agent': USER_AGENT})
                soup = BeautifulSoup(res.text, 'lxml')
                price = soup.find('span', {'class': 'price-standard'}).text.strip().lstrip('€ ').replace(',', '.')
                sale_price = soup.find('span', {'data-ta-id': 'productvariantcontentPrice'}).text.strip().lstrip('€ ').replace(',', '.')

                price = self.get_cos_price(price)
                sale_price = self.get_cos_price(sale_price)

                color = soup.find('b').text
                if not size:  # сумки
                    pass
                else:
                    sizes = soup.find_all('button', {'class': 'swatch--size'})
                    if size:
                        for size_elem in sizes:
                            new_size = size_elem.text.strip()
                            if length:  # брюки
                                lengths = soup.find_all('button', {'class': '.swatch--length'})
                                for length_elem in lengths:
                                    new_length = length_elem.text.strip()
                                    new_article = prefix + '_' + article + '_' + color + '_' + new_size + '_' + new_length
                                    if 'swatch--noStock' in length_elem.get('class') or 'swatch--noStock' in new_size.get('class'):
                                        self.result[new_article] = [self.AVIABLE_UNIQLO['no_aviable'], price, sale_price]
                                    else:
                                        self.result[new_article] = [self.AVIABLE_UNIQLO['aviable'], price, sale_price]
                            else:  # остальное
                                new_article = prefix + '_' + article + '_' + color + '_' + new_size
                                if 'swatch--noStock' in size_elem.get('class'):
                                    self.result[new_article] = [self.AVIABLE_UNIQLO['no_aviable'], price, sale_price]
                                else:
                                    self.result[new_article] = [self.AVIABLE_UNIQLO['aviable'], price, sale_price]
                    else:  # сумки
                        new_article = prefix + '_' + article + '_' + color
                        aviable = soup.find('button', {'class': 'swatch--size'})
                        if 'swatch--noStock' in aviable.get('class'):
                            self.result[new_article] = [self.AVIABLE_UNIQLO['no_aviable'], price, sale_price]
                        else:
                            self.result[new_article] = [self.AVIABLE_UNIQLO['aviable'], price, sale_price]

    def parse(self, articles):
        parsed_articles = []
        parse_brands = ['H&M', 'COS', 'UNIQLO', 'ASOS']
        for i in articles:
            print(f'{articles.index(i) + 1} of {len(articles)}. Article: {i}')

            parts_of_article = i.split('_')
            if parts_of_article[0] == 'UNIQLO':
                if len(parts_of_article) == 4:
                    prefix, article, color, size = parts_of_article
                    length = None
                elif len(parts_of_article) == 5:
                    prefix, article, color, size, length = parts_of_article
                else:
                    prefix, article, color = parts_of_article
                    size, length = None, None
            else:
                if len(parts_of_article) == 3:
                    prefix, article, size = parts_of_article
                    length = None
                elif len(parts_of_article) > 3:
                    prefix, article, size, length = parts_of_article[0], parts_of_article[1], '_'.join(parts_of_article[2]), parts_of_article[3]
                elif len(parts_of_article) == 2:
                    prefix, article = parts_of_article
                    size, length = None, None
                else:
                    print(i)
                    raise Exception('Article has less than 3 parts')
            if (prefix not in parse_brands) or i in parsed_articles:
                continue

            parsed_articles.append(i)
            try:
                self.parseOne(prefix, article, size, length)
            except TimeoutException:
                pass
            except Exception:
                self.driver.refresh()
                self.writeLog()

        return self.result

    def gPriceDict(self, key):
        return float(PRICE_TABLE[key])

    def get_hm_price(self, pln_price):
        cost_price = ((float(pln_price) / self.gPriceDict("КУРС_USD_ЗЛОТЫ")) * self.gPriceDict("КОЭФ_КОНВЕРТАЦИИ") * self.gPriceDict(
            'КУРС_USD_RUB')) + (self.gPriceDict('ЦЕНА_ДОСТАВКИ_В_КАТЕГОРИИ') * self.gPriceDict('КУРС_БЕЛ.РУБ_РУБ') * self.gPriceDict(
            'КУРС_EUR_БЕЛ.РУБ'))
        final_price = (cost_price + self.gPriceDict('СРЕД_ЦЕН_ДОСТАВКИ')) / (
                    1 - self.gPriceDict('НАЦЕНКА') - self.gPriceDict('ПРОЦЕНТЫ_ОЗОН') - self.gPriceDict('ПРОЦЕНТЫ_НАЛОГ') - self.gPriceDict('ПРОЦЕНТЫ_ЭКВАЙРИНГ'))

        final_price = (final_price // 100 + 1) * 100 - 1
        return final_price

    def get_cos_price(self, eur_price):
        cost_price = (float(eur_price) * self.gPriceDict("КОЭФ_КОНВЕРТАЦИИ") * self.gPriceDict(
            'КУРС_EUR_RUB')) + (self.gPriceDict("ЦЕНА_ДОСТАВКИ_В_КАТЕГОРИИ") * self.gPriceDict(
            'КУРС_БЕЛ.РУБ_РУБ') * self.gPriceDict(
            'КУРС_EUR_БЕЛ.РУБ'))
        final_price = (cost_price + self.gPriceDict('СРЕД_ЦЕН_ДОСТАВКИ')) / (
                1 - self.gPriceDict('НАЦЕНКА') - self.gPriceDict("ПРОЦЕНТЫ_ОЗОН") - self.gPriceDict(
            'ПРОЦЕНТЫ_НАЛОГ') - self.gPriceDict('ПРОЦЕНТЫ_ЭКВАЙРИНГ'))

        final_price = (final_price // 100 + 1) * 100 - 10
        return final_price

    def make_request(self, url, method='get', headers=None, cookies=None, files=None, retries=3, delay=1):
        urllib3.disable_warnings(InsecureRequestWarning) # убираем предупреждения для невалидных https сертов
        
        for _ in range(retries + 1):
            try:
                if method == 'post':
                    response = requests.post(url, headers=headers, cookies=cookies, files=files, verify=False)
                else:
                    response = requests.get(url, headers=headers, cookies=cookies)
                response.raise_for_status()  # Бросает исключение для 4xx и 5xx кодов ответа
                return response
            except RequestException as e:
                print(f"Error: {e}")
                if _ < retries:
                    print(f"Retrying in {delay} seconds...")
                    sleep(delay)
        return None  # Если все попытки завершились неудачей

    def save(self, result):
        wb = load_workbook(filename=f'templates/{TEMPLATE_NAME}')
        ws = wb['Остатки на складе']

        data_validation = DataValidation(
            type="list",
            formula1='',  # Значения списка
            showDropDown=True  # Отображать выпадающий список в ячейке
        )

        for i in range(2, ws.max_row + 1):
            if ws['B' + str(i)].value:  # Проверка на пустую строку
                if ws['B' + str(i)].value[:3] == 'COS':
                    data_validation.add("A" + str(i))
                    ws.cell(row=i, column=1).value = self.AVIABLE_COS['stock_name']
                elif ws['B' + str(i)].value[:3] == 'H&M':
                    data_validation.add("A" + str(i))
                    ws.cell(row=i, column=1).value = self.AVIABLE_HM['stock_name']
                elif ws['B' + str(i)].value[:6] == 'UNIQLO':
                    data_validation.add("A" + str(i))
                    ws.cell(row=i, column=1).value = self.AVIABLE_UNIQLO['stock_name']
                elif ws['B' + str(i)].value[:4] == 'ASOS':
                    data_validation.add("A" + str(i))
                    ws.cell(row=i, column=1).value = self.AVIABLE_ASOS['stock_name']
                try:
                    ws.cell(row=i, column=4).value = result[ws['B' + str(i)].value][0]
                    ws.cell(row=i, column=6).value = result[ws['B' + str(i)].value][1]
                    ws.cell(row=i, column=7).value = result[ws['B' + str(i)].value][2]
                except KeyError:
                    pass

        wb.save(SAVE_XLSX_PATH + f"{datetime.now()}.xlsx".replace(':', '.'))

    def get_articles(self):
        articles = []

        wb = load_workbook(filename=f'templates/{TEMPLATE_NAME}')
        ws = wb['Остатки на складе']
        for i in range(2, ws.max_row + 1):
            articles.append(ws['B' + str(i)].value)

        return articles

    def get_urls_sizes(self, articles):
        urls_sizes = {}
        for article in articles:
            prefix, article, size = article.split('_')
            if prefix == 'H&M':
                url = f'https://www2.hm.com/pl_pl/productpage.{article}.html'
                if url in urls_sizes.keys():
                    urls_sizes[url] += [size]
                else:
                    urls_sizes[url] = [size]
        return urls_sizes

    def check_exists_by_xpath(self, elem, xpath):
        try:
            elem.find_element(By.XPATH, xpath)
        except NoSuchElementException:
            return False
        return True

    def delete_duplicates(self, articles):
        result = []
        tmp = []
        for article in articles:
            if article and article.split('_')[1] not in tmp:
                tmp.append(article.split('_')[1])
                result.append(article)
        return result

    def load_settings(self):
        with open('config/aviable.json', 'r', encoding='utf-8') as f:
            self.settings = json.load(f)
        self.AVIABLE_HM = self.settings['H&M']
        self.AVIABLE_COS = self.settings['COS']
        self.AVIABLE_UNIQLO = self.settings['UNIQLO']
        self.AVIABLE_ASOS = self.settings['ASOS']

    def start(self):
        try:
            print('--- START PARSING ---')
            self.load_settings()
            articles = self.delete_duplicates(self.get_articles())
            result = self.parse(articles)
            self.save(result)
            print('--- END PARSING ---')
        except:
            self.writeLog()
            with open('last.html', 'w') as f:
                f.write(self.driver.page_source)
        finally:
            self.driver.close()
            self.driver.quit()

    def writeLog(self):
        error = self.driver.current_url + '\n' + traceback.format_exc() + '\n'
        print(error)
        with open('log.log', 'a') as f:
            f.write(error)


def main():
    parser = Parser()
    parser.start()


if __name__ == '__main__':
    if 'xlsx' not in os.listdir():
        os.mkdir('xlsx')
    main()
